import csv
import sqlite3
import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "python"))

from generate_data import generate
from run_pipeline import run


class PipelineTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        (self.root / "docs").mkdir()
        generate(self.root / "data" / "raw", seed=11, prospect_count=700)
        run(self.root)
        self.conn = sqlite3.connect(self.root / "data" / "processed" / "analytics.db")

    def tearDown(self):
        self.conn.close()
        self.temp.cleanup()

    def test_unique_primary_keys(self):
        for table, key in [("dim_prospect", "prospect_id"), ("dim_customer", "customer_id"), ("fct_orders", "order_id")]:
            total, unique = self.conn.execute(f"SELECT COUNT(*), COUNT(DISTINCT {key}) FROM {table}").fetchone()
            self.assertEqual(total, unique)

    def test_foreign_keys_are_valid(self):
        self.assertEqual(self.conn.execute("PRAGMA foreign_key_check").fetchall(), [])

    def test_orders_reconcile_to_channel_mart(self):
        raw_total = self.conn.execute("SELECT ROUND(SUM(contribution_margin),2) FROM fct_orders").fetchone()[0]
        mart_total = self.conn.execute("SELECT ROUND(SUM(contribution_margin),2) FROM mart_channel_profitability").fetchone()[0]
        self.assertAlmostEqual(raw_total, mart_total, places=2)

    def test_experiment_has_converters_and_non_converters(self):
        minimum, maximum = self.conn.execute("SELECT MIN(converted), MAX(converted) FROM dim_prospect").fetchone()
        self.assertEqual((minimum, maximum), (0, 1))
        rates = [row[0] for row in self.conn.execute("SELECT conversion_rate FROM mart_experiment")]
        self.assertTrue(all(0 < rate < 1 for rate in rates))

    def test_experiment_output_has_statistics(self):
        with (self.root / "data" / "processed" / "experiment.csv").open(newline="", encoding="utf-8") as handle:
            rows = list(csv.DictReader(handle))
        self.assertIn("conversion_p_value", rows[0])
        self.assertIn("srm_p_value", rows[0])

    def test_typed_schema(self):
        columns = {row[1]: row[2] for row in self.conn.execute("PRAGMA table_info(fct_orders)")}
        self.assertEqual(columns["gross_revenue"], "REAL")
        self.assertEqual(columns["order_id"], "TEXT")

    def test_artifacts_created(self):
        self.assertTrue((self.root / "artifacts" / "Baltic_Commerce_Analysis.xlsx").exists())
        self.assertTrue((self.root / "docs" / "index.html").exists())

    def test_excel_is_real_workbook_with_recruiter_ready_sheets(self):
        from openpyxl import load_workbook

        workbook = load_workbook(self.root / "artifacts" / "Baltic_Commerce_Analysis.xlsx", data_only=False)
        self.assertEqual(
            workbook.sheetnames,
            ["Executive Overview", "Channel Analysis", "Market Analysis", "Logistics Analysis", "Experiment Analysis", "Data Dictionary"],
        )
        self.assertTrue(workbook["Channel Analysis"]["I2"].value.startswith("=IF("))
        self.assertGreater(len(workbook["Channel Analysis"]._charts), 0)

    def test_dashboard_has_no_negative_svg_widths(self):
        dashboard = (self.root / "docs" / "index.html").read_text(encoding="utf-8")
        self.assertNotIn('width="-', dashboard)


if __name__ == "__main__":
    unittest.main()
