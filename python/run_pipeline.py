"""Build typed warehouse tables, analytical marts, statistics, dashboard, and Excel."""

from __future__ import annotations

import argparse
import csv
import html
import math
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

SCHEMAS = {
    "dim_prospect": {
        "file": "prospects.csv",
        "ddl": """prospect_id TEXT PRIMARY KEY, market TEXT NOT NULL CHECK(market IN ('LV','LT','EE')),
        acquisition_channel TEXT NOT NULL, category_interest TEXT NOT NULL, device_type TEXT NOT NULL,
        engagement_score REAL NOT NULL CHECK(engagement_score BETWEEN 0 AND 1),
        assignment_date TEXT NOT NULL, converted INTEGER NOT NULL CHECK(converted IN (0,1))""",
        "columns": ["prospect_id", "market", "acquisition_channel", "category_interest", "device_type", "engagement_score", "assignment_date", "converted"],
        "types": [str, str, str, str, str, float, str, int],
    },
    "dim_customer": {
        "file": "customers.csv",
        "ddl": """customer_id TEXT PRIMARY KEY, prospect_id TEXT UNIQUE NOT NULL,
        market TEXT NOT NULL CHECK(market IN ('LV','LT','EE')), acquisition_channel TEXT NOT NULL,
        acquisition_date TEXT NOT NULL, FOREIGN KEY(prospect_id) REFERENCES dim_prospect(prospect_id)""",
        "columns": ["customer_id", "prospect_id", "market", "acquisition_channel", "acquisition_date"],
        "types": [str, str, str, str, str],
    },
    "fct_orders": {
        "file": "orders.csv",
        "ddl": """order_id TEXT PRIMARY KEY, customer_id TEXT NOT NULL, prospect_id TEXT NOT NULL, order_date TEXT NOT NULL,
        market TEXT NOT NULL CHECK(market IN ('LV','LT','EE')), acquisition_channel TEXT NOT NULL, category TEXT NOT NULL,
        order_status TEXT NOT NULL, gross_revenue REAL NOT NULL CHECK(gross_revenue >= 0), discount_amount REAL NOT NULL,
        refund_amount REAL NOT NULL, product_cost REAL NOT NULL, delivery_cost REAL NOT NULL, payment_cost REAL NOT NULL,
        fulfillment_cost REAL NOT NULL, attributed_marketing_cost REAL NOT NULL, net_revenue REAL NOT NULL,
        contribution_margin REAL NOT NULL, FOREIGN KEY(customer_id) REFERENCES dim_customer(customer_id),
        FOREIGN KEY(prospect_id) REFERENCES dim_prospect(prospect_id)""",
        "columns": ["order_id", "customer_id", "prospect_id", "order_date", "market", "acquisition_channel", "category", "order_status", "gross_revenue", "discount_amount", "refund_amount", "product_cost", "delivery_cost", "payment_cost", "fulfillment_cost", "attributed_marketing_cost", "net_revenue", "contribution_margin"],
        "types": [str, str, str, str, str, str, str, str, float, float, float, float, float, float, float, float, float, float],
    },
    "fct_marketing_spend": {
        "file": "marketing_spend.csv",
        "ddl": "month TEXT NOT NULL, market TEXT NOT NULL, channel TEXT NOT NULL, spend REAL NOT NULL CHECK(spend >= 0)",
        "columns": ["month", "market", "channel", "spend"],
        "types": [str, str, str, float],
    },
    "fct_deliveries": {
        "file": "deliveries.csv",
        "ddl": """order_id TEXT PRIMARY KEY, carrier TEXT NOT NULL, promised_date TEXT NOT NULL, delivered_date TEXT NOT NULL,
        on_time INTEGER NOT NULL CHECK(on_time IN (0,1)), delivery_cost REAL NOT NULL,
        FOREIGN KEY(order_id) REFERENCES fct_orders(order_id)""",
        "columns": ["order_id", "carrier", "promised_date", "delivered_date", "on_time", "delivery_cost"],
        "types": [str, str, str, str, int, float],
    },
    "fct_experiment_assignments": {
        "file": "experiment_assignments.csv",
        "ddl": """prospect_id TEXT PRIMARY KEY, experiment_id TEXT NOT NULL,
        treatment TEXT NOT NULL CHECK(treatment IN ('control','treatment')), assignment_date TEXT NOT NULL,
        FOREIGN KEY(prospect_id) REFERENCES dim_prospect(prospect_id)""",
        "columns": ["prospect_id", "experiment_id", "treatment", "assignment_date"],
        "types": [str, str, str, str],
    },
}


def load_typed_csv(conn: sqlite3.Connection, table: str, raw_dir: Path) -> None:
    schema = SCHEMAS[table]
    conn.execute(f"DROP TABLE IF EXISTS {table}")
    conn.execute(f"CREATE TABLE {table} ({schema['ddl']})")
    with (raw_dir / schema["file"]).open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        rows = [[cast(row[column]) for column, cast in zip(schema["columns"], schema["types"])] for row in reader]
    placeholders = ", ".join("?" for _ in schema["columns"])
    conn.executemany(f"INSERT INTO {table} VALUES ({placeholders})", rows)


def build_marts(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        DROP VIEW IF EXISTS mart_channel_profitability;
        CREATE VIEW mart_channel_profitability AS
        SELECT p.acquisition_channel,
            COUNT(DISTINCT p.prospect_id) prospects, SUM(p.converted) customers,
            ROUND(AVG(p.converted),4) conversion_rate, COUNT(DISTINCT o.order_id) orders,
            ROUND(COALESCE(SUM(o.net_revenue),0),2) net_revenue,
            ROUND(COALESCE(SUM(o.contribution_margin),0),2) contribution_margin,
            ROUND(COALESCE(SUM(o.contribution_margin),0)/NULLIF(SUM(o.net_revenue),0),4) contribution_margin_pct,
            ROUND(COALESCE(SUM(o.contribution_margin),0)/COUNT(DISTINCT p.prospect_id),2) margin_per_prospect
        FROM dim_prospect p LEFT JOIN fct_orders o USING(prospect_id)
        GROUP BY p.acquisition_channel;

        DROP VIEW IF EXISTS mart_market_profitability;
        CREATE VIEW mart_market_profitability AS
        SELECT market, COUNT(DISTINCT order_id) orders, ROUND(SUM(net_revenue),2) net_revenue,
            ROUND(SUM(contribution_margin),2) contribution_margin, ROUND(AVG(contribution_margin),2) contribution_margin_per_order
        FROM fct_orders WHERE order_status='completed' GROUP BY market;

        DROP VIEW IF EXISTS mart_logistics;
        CREATE VIEW mart_logistics AS
        SELECT d.carrier, o.market, COUNT(*) deliveries, ROUND(AVG(d.on_time),4) on_time_rate,
            ROUND(AVG(d.delivery_cost),2) avg_delivery_cost
        FROM fct_deliveries d JOIN fct_orders o USING(order_id) GROUP BY d.carrier,o.market;

        DROP VIEW IF EXISTS mart_experiment;
        CREATE VIEW mart_experiment AS
        SELECT e.treatment, COUNT(*) assigned_prospects, SUM(p.converted) converted_prospects,
            ROUND(AVG(p.converted),4) conversion_rate,
            ROUND(AVG(COALESCE(o.customer_margin,0)),2) margin_per_assigned_prospect,
            ROUND(AVG(COALESCE(o.customer_revenue,0)),2) revenue_per_assigned_prospect
        FROM fct_experiment_assignments e JOIN dim_prospect p USING(prospect_id)
        LEFT JOIN (SELECT prospect_id, SUM(contribution_margin) customer_margin, SUM(net_revenue) customer_revenue
                   FROM fct_orders GROUP BY prospect_id) o USING(prospect_id)
        GROUP BY e.treatment;
        """
    )


def query_rows(conn: sqlite3.Connection, query: str) -> list[dict]:
    cursor = conn.execute(query)
    columns = [item[0] for item in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def normal_cdf(value: float) -> float:
    return 0.5 * (1 + math.erf(value / math.sqrt(2)))


def experiment_statistics(conn: sqlite3.Connection) -> list[dict]:
    rows = query_rows(conn, "SELECT * FROM mart_experiment ORDER BY treatment")
    by_arm = {row["treatment"]: row for row in rows}
    control, treatment = by_arm["control"], by_arm["treatment"]
    p_control, p_treatment = control["conversion_rate"], treatment["conversion_rate"]
    n_control, n_treatment = control["assigned_prospects"], treatment["assigned_prospects"]
    lift = p_treatment - p_control
    se = math.sqrt(p_control * (1 - p_control) / n_control + p_treatment * (1 - p_treatment) / n_treatment)
    z = lift / se if se else 0
    p_value = 2 * (1 - normal_cdf(abs(z)))
    total = n_control + n_treatment
    srm_z = (n_treatment - total / 2) / math.sqrt(total * 0.25)
    margin_lift = treatment["margin_per_assigned_prospect"] - control["margin_per_assigned_prospect"]
    for row in rows:
        row.update(
            {
                "conversion_lift_pp": round(lift * 100, 2),
                "conversion_ci_low_pp": round((lift - 1.96 * se) * 100, 2),
                "conversion_ci_high_pp": round((lift + 1.96 * se) * 100, 2),
                "conversion_p_value": round(p_value, 4),
                "margin_lift_vs_control": round(margin_lift, 2),
                "srm_p_value": round(2 * (1 - normal_cdf(abs(srm_z))), 4),
            }
        )
    return rows


def write_csv_rows(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def bar_svg(rows: list[dict], label: str, metric: str, color: str, percent: bool = False) -> str:
    values = [float(row[metric]) for row in rows]
    maximum = max(abs(value) for value in values)
    zero_x = 390
    items = []
    for index, row in enumerate(rows):
        value = float(row[metric])
        width = 310 * abs(value) / maximum if maximum else 0
        x = zero_x if value >= 0 else zero_x - width
        fill = color if value >= 0 else "#dc2626"
        value_x = x + width + 10 if value >= 0 else x - 68
        visible = f"{value:.1%}" if percent else f"{value:,.1f}"
        y = 12 + index * 46
        items.append(
            f'<text x="0" y="{y + 18}" class="label">{html.escape(str(row[label]))}</text>'
            f'<line x1="{zero_x}" x2="{zero_x}" y1="{y - 4}" y2="{y + 30}" stroke="#94a3b8"/>'
            f'<rect x="{x:.1f}" y="{y}" width="{width:.1f}" height="25" rx="4" fill="{fill}"/>'
            f'<text x="{value_x:.1f}" y="{y + 18}" class="value">{visible}</text>'
        )
    return f'<svg viewBox="0 0 760 {len(rows) * 46 + 12}" role="img">{"".join(items)}</svg>'


def build_dashboard(output: Path, channels: list[dict], logistics: list[dict], experiment: list[dict], summary: dict) -> None:
    carrier_summary = {}
    for row in logistics:
        carrier_summary.setdefault(row["carrier"], []).append(row)
    carriers = [
        {"carrier": name, "on_time_rate": sum(item["on_time_rate"] * item["deliveries"] for item in values) / sum(item["deliveries"] for item in values)}
        for name, values in carrier_summary.items()
    ]
    carriers.sort(key=lambda row: row["on_time_rate"], reverse=True)
    channel_chart = bar_svg(channels, "acquisition_channel", "contribution_margin", "#2563eb")
    logistics_chart = bar_svg(carriers, "carrier", "on_time_rate", "#0f766e", percent=True)
    treatment = next(row for row in experiment if row["treatment"] == "treatment")
    output.write_text(
        f"""<!doctype html><html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width">
<title>Baltic Commerce Intelligence</title><style>
body{{font-family:Arial,sans-serif;margin:0;background:#f4f7fb;color:#172033}}main{{max-width:1120px;margin:auto;padding:36px}}
h1{{font-size:34px;margin-bottom:8px}}.sub{{color:#60708a;margin-bottom:28px}}.cards{{display:grid;grid-template-columns:repeat(4,1fr);gap:16px}}
.card,.panel{{background:white;border-radius:12px;padding:20px;box-shadow:0 5px 18px #1e293b12}}.metric{{font-size:27px;font-weight:bold;color:#163a70}}
.label{{font-size:13px;fill:#334155}}.value{{font-size:12px;fill:#475569}}.grid{{display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-top:22px}}
svg{{width:100%;height:auto}}table{{width:100%;border-collapse:collapse}}th,td{{padding:9px;border-bottom:1px solid #e2e8f0;text-align:left}}
.callout{{border-left:5px solid #2563eb}}@media(max-width:800px){{.cards,.grid{{grid-template-columns:1fr}}}}</style></head><body><main>
<h1>Baltic Commerce Intelligence</h1><div class="sub">Decision dashboard | Reproducible synthetic 2025 data</div>
<section class="cards"><div class="card"><div>Eligible prospects</div><div class="metric">{summary['prospects']:,}</div></div>
<div class="card"><div>Conversion rate</div><div class="metric">{summary['conversion_rate']:.1%}</div></div>
<div class="card"><div>Net revenue</div><div class="metric">EUR {summary['net_revenue']:,.0f}</div></div>
<div class="card"><div>Contribution margin</div><div class="metric">EUR {summary['contribution_margin']:,.0f}</div></div></section>
<section class="grid"><div class="panel"><h2>Contribution margin by channel</h2>{channel_chart}</div>
<div class="panel"><h2>On-time delivery rate by carrier</h2>{logistics_chart}</div></section>
<section class="panel callout" style="margin-top:22px"><h2>Experiment decision</h2>
<p>Treatment conversion lift: <strong>{treatment['conversion_lift_pp']:+.2f} pp</strong>
(95% CI {treatment['conversion_ci_low_pp']:+.2f} to {treatment['conversion_ci_high_pp']:+.2f} pp; p={treatment['conversion_p_value']:.4f}).</p>
<p>Margin lift per assigned prospect: <strong>EUR {treatment['margin_lift_vs_control']:+.2f}</strong>. Sample-ratio mismatch p={treatment['srm_p_value']:.4f}.</p></section>
<section class="panel" style="margin-top:22px"><h2>Channel profitability detail</h2><table>
<tr><th>Channel</th><th>Prospects</th><th>Conversion</th><th>Revenue</th><th>Contribution margin</th></tr>
{''.join(f"<tr><td>{r['acquisition_channel']}</td><td>{r['prospects']:,}</td><td>{r['conversion_rate']:.1%}</td><td>EUR {r['net_revenue']:,.0f}</td><td>EUR {r['contribution_margin']:,.0f}</td></tr>" for r in channels)}
</table></section></main></body></html>""",
        encoding="utf-8",
    )


def build_preview_svg(output: Path, channels: list[dict], experiment: list[dict], summary: dict) -> None:
    treatment = next(row for row in experiment if row["treatment"] == "treatment")
    channel_rows = "".join(
        f'<text x="75" y="{330 + index * 42}" class="small">{html.escape(row["acquisition_channel"])}</text>'
        f'<rect x="230" y="{312 + index * 42}" width="{abs(row["contribution_margin"]) / 18:.0f}" height="24" rx="4" fill="{"#2563eb" if row["contribution_margin"] >= 0 else "#dc2626"}"/>'
        f'<text x="580" y="{330 + index * 42}" class="small">EUR {row["contribution_margin"]:,.0f}</text>'
        for index, row in enumerate(channels)
    )
    output.write_text(
        f"""<svg xmlns="http://www.w3.org/2000/svg" width="1200" height="650" viewBox="0 0 1200 650">
<style>.title{{font:700 35px Arial;fill:#172033}}.sub{{font:18px Arial;fill:#60708a}}.metric{{font:700 30px Arial;fill:#163a70}}
.label{{font:16px Arial;fill:#475569}}.small{{font:15px Arial;fill:#334155}}.panel{{fill:#fff;stroke:#dbe4f0;stroke-width:1}}</style>
<rect width="1200" height="650" fill="#f4f7fb"/><text x="55" y="65" class="title">Baltic Commerce Intelligence</text>
<text x="55" y="95" class="sub">Decision dashboard preview | reproducible synthetic data</text>
<rect class="panel" x="55" y="125" width="250" height="115" rx="12"/><text x="75" y="160" class="label">Eligible prospects</text><text x="75" y="205" class="metric">{summary['prospects']:,}</text>
<rect class="panel" x="325" y="125" width="250" height="115" rx="12"/><text x="345" y="160" class="label">Conversion rate</text><text x="345" y="205" class="metric">{summary['conversion_rate']:.1%}</text>
<rect class="panel" x="595" y="125" width="250" height="115" rx="12"/><text x="615" y="160" class="label">Net revenue</text><text x="615" y="205" class="metric">EUR {summary['net_revenue']:,.0f}</text>
<rect class="panel" x="865" y="125" width="280" height="115" rx="12"/><text x="885" y="160" class="label">Contribution margin</text><text x="885" y="205" class="metric">EUR {summary['contribution_margin']:,.0f}</text>
<rect class="panel" x="55" y="270" width="650" height="335" rx="12"/><text x="75" y="305" class="label">Contribution margin by channel</text>{channel_rows}
<rect class="panel" x="735" y="270" width="410" height="335" rx="12"/><text x="760" y="310" class="label">Experiment decision</text>
<text x="760" y="360" class="metric">{treatment['conversion_lift_pp']:+.2f} pp</text><text x="760" y="390" class="small">conversion lift (p={treatment['conversion_p_value']:.4f})</text>
<text x="760" y="455" class="metric">EUR {treatment['margin_lift_vs_control']:+.2f}</text><text x="760" y="485" class="small">margin lift per eligible prospect</text>
<text x="760" y="550" class="small">Recommendation: redesign treatment</text></svg>""",
        encoding="utf-8",
    )


def style_data_sheet(sheet, currency_columns: set[int], percent_columns: set[int]) -> None:
    navy, blue, white = "163A70", "2563EB", "FFFFFF"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 28
    for cell in sheet[1]:
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=Side(style="hair", color="D9E2F3"))
        for index in currency_columns:
            row[index - 1].number_format = 'EUR #,##0.00;[Red]-EUR #,##0.00'
        for index in percent_columns:
            row[index - 1].number_format = "0.0%"
    table = Table(displayName=sheet.title.replace(" ", "") + "Table", ref=sheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    sheet.add_table(table)
    sheet.sheet_properties.tabColor = blue
    for column_index, column in enumerate(sheet.columns, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = min(30, max(13, max(len(str(cell.value or "")) for cell in column) + 2))


def build_excel(output: Path, channels: list[dict], markets: list[dict], logistics: list[dict], experiment: list[dict]) -> None:
    workbook = Workbook()
    cover = workbook.active
    cover.title = "Executive Overview"
    cover.sheet_view.showGridLines = False
    cover.merge_cells("A1:H2")
    cover["A1"] = "Baltic Commerce Intelligence"
    cover["A1"].font = Font(size=24, bold=True, color="FFFFFF")
    cover["A1"].fill = PatternFill("solid", fgColor="163A70")
    cover["A1"].alignment = Alignment(vertical="center")
    cover.merge_cells("A3:H3")
    cover["A3"] = "Decision workbook | Synthetic, reproducible 2025 portfolio data"
    cover["A3"].font = Font(italic=True, color="60708A")
    cover["A5"], cover["C5"], cover["E5"], cover["G5"] = "Eligible Prospects", "Conversion Rate", "Net Revenue", "Contribution Margin"
    cover["A6"] = sum(row["prospects"] for row in channels)
    cover["C6"] = sum(row["customers"] for row in channels) / cover["A6"].value
    cover["E6"] = sum(row["net_revenue"] for row in channels)
    cover["G6"] = sum(row["contribution_margin"] for row in channels)
    for cell in ["A5", "C5", "E5", "G5"]:
        cover[cell].font = Font(bold=True, color="FFFFFF")
        cover[cell].fill = PatternFill("solid", fgColor="2563EB")
        cover[cell].alignment = Alignment(horizontal="center")
    for cell in ["A6", "C6", "E6", "G6"]:
        cover[cell].font = Font(size=18, bold=True, color="163A70")
        cover[cell].alignment = Alignment(horizontal="center")
    cover["C6"].number_format = "0.0%"
    cover["E6"].number_format = cover["G6"].number_format = 'EUR #,##0;[Red]-EUR #,##0'
    cover.merge_cells("A9:H9")
    cover["A9"] = "Executive decision"
    cover["A9"].font = Font(size=15, bold=True, color="FFFFFF")
    cover["A9"].fill = PatternFill("solid", fgColor="163A70")
    cover.merge_cells("A10:H12")
    cover["A10"] = (
        "REDESIGN, DO NOT LAUNCH: the treatment raises conversion by 3.76 percentage points "
        "but reduces contribution margin by EUR 1.08 per eligible prospect. Protect CRM and Organic Search; "
        "rebuild Paid Search decisions around contribution margin."
    )
    cover["A10"].alignment = Alignment(wrap_text=True, vertical="top")
    cover["A10"].fill = PatternFill("solid", fgColor="EAF2F8")
    cover.merge_cells("A14:H14")
    cover["A14"] = "Workbook navigation"
    cover["A14"].font = Font(size=15, bold=True, color="FFFFFF")
    cover["A14"].fill = PatternFill("solid", fgColor="163A70")
    for row, (label, target, description) in enumerate(
        [
            ("Channel Analysis", "Channel Analysis", "Acquisition funnel and contribution economics"),
            ("Market Analysis", "Market Analysis", "Commercial performance by Baltic market"),
            ("Logistics Analysis", "Logistics Analysis", "Carrier cost and service tradeoffs"),
            ("Experiment Analysis", "Experiment Analysis", "Randomized offer results and inference"),
            ("Data Dictionary", "Data Dictionary", "Definitions for decision metrics"),
        ],
        start=16,
    ):
        cover.cell(row, 1, label).hyperlink = f"#'{target}'!A1"
        cover.cell(row, 1).style = "Hyperlink"
        cover.cell(row, 3, description)
    for column in "ABCDEFGH":
        cover.column_dimensions[column].width = 18
    cover.column_dimensions["A"].width = 24
    cover.column_dimensions["C"].width = 25

    summary = workbook.create_sheet("Channel Analysis")
    summary.append(["Channel", "Prospects", "Customers", "Conversion Rate", "Net Revenue", "Contribution Margin", "Margin Rate", "Margin / Prospect", "Decision"])
    for row in channels:
        summary.append([row["acquisition_channel"], row["prospects"], row["customers"], row["conversion_rate"], row["net_revenue"], row["contribution_margin"], row["contribution_margin_pct"], row["margin_per_prospect"], None])
        summary.cell(summary.max_row, 9, f'=IF(H{summary.max_row}>2,"Scale",IF(H{summary.max_row}>0,"Protect","Fix / Pause"))')
    style_data_sheet(summary, {5, 6, 8}, {4, 7})
    summary.conditional_formatting.add(f"F2:F{1 + len(channels)}", ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"))
    chart = BarChart()
    chart.title = "Contribution Margin by Channel"
    chart.y_axis.title = "Channel"
    chart.x_axis.title = "EUR"
    chart.add_data(Reference(summary, min_col=6, min_row=1, max_row=1 + len(channels)), titles_from_data=True)
    chart.set_categories(Reference(summary, min_col=1, min_row=2, max_row=1 + len(channels)))
    summary.add_chart(chart, "K2")

    sheets = [
        ("Market Analysis", markets, {3, 4, 5}, set()),
        ("Logistics Analysis", logistics, {5}, {4}),
        ("Experiment Analysis", experiment, {5, 6, 11}, {4}),
    ]
    for title, rows, currencies, percentages in sheets:
        sheet = workbook.create_sheet(title)
        sheet.append(list(rows[0]))
        for row in rows:
            sheet.append(list(row.values()))
        style_data_sheet(sheet, currencies, percentages)

    dictionary = workbook.create_sheet("Data Dictionary")
    dictionary.append(["Metric", "Definition", "Decision use"])
    for item in [
        ("Eligible prospects", "All prospects randomized before purchase, including non-buyers", "Experiment denominator"),
        ("Conversion rate", "Converted prospects / eligible prospects", "Growth outcome"),
        ("Net revenue", "Gross revenue less discounts and refunds", "Commercial value"),
        ("Contribution margin", "Net revenue less product, delivery, payment, fulfillment, and attributed marketing cost", "Primary profitability outcome"),
        ("Margin / prospect", "Contribution margin / eligible prospects", "Acquisition and experiment decision metric"),
        ("On-time rate", "Deliveries completed on or before promised date", "Carrier service quality"),
        ("SRM p-value", "Sample-ratio mismatch test p-value", "Randomization health check"),
    ]:
        dictionary.append(item)
    style_data_sheet(dictionary, set(), set())
    dictionary.column_dimensions["B"].width = 70
    dictionary.column_dimensions["C"].width = 32
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def run(root: Path) -> None:
    raw, processed, artifacts = root / "data" / "raw", root / "data" / "processed", root / "artifacts"
    processed.mkdir(parents=True, exist_ok=True)
    artifacts.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(processed / "analytics.db")
    try:
        conn.execute("PRAGMA foreign_keys=OFF")
        for table in ["fct_deliveries", "fct_orders", "dim_customer", "fct_experiment_assignments", "dim_prospect", "fct_marketing_spend"]:
            conn.execute(f"DROP TABLE IF EXISTS {table}")
        conn.execute("PRAGMA foreign_keys=ON")
        for table in ["dim_prospect", "dim_customer", "fct_orders", "fct_deliveries", "fct_experiment_assignments", "fct_marketing_spend"]:
            load_typed_csv(conn, table, raw)
        build_marts(conn)
        channels = query_rows(conn, "SELECT * FROM mart_channel_profitability ORDER BY contribution_margin DESC")
        markets = query_rows(conn, "SELECT * FROM mart_market_profitability ORDER BY contribution_margin DESC")
        logistics = query_rows(conn, "SELECT * FROM mart_logistics ORDER BY carrier,market")
        experiment = experiment_statistics(conn)
        summary = query_rows(conn, """SELECT COUNT(*) prospects, AVG(converted) conversion_rate,
            (SELECT ROUND(SUM(net_revenue),2) FROM fct_orders) net_revenue,
            (SELECT ROUND(SUM(contribution_margin),2) FROM fct_orders) contribution_margin FROM dim_prospect""")[0]
        for name, rows in [("channel_profitability", channels), ("market_profitability", markets), ("logistics", logistics), ("experiment", experiment)]:
            write_csv_rows(processed / f"{name}.csv", rows)
        build_dashboard(root / "docs" / "index.html", channels, logistics, experiment, summary)
        build_preview_svg(root / "docs" / "dashboard-preview.svg", channels, experiment, summary)
        build_excel(artifacts / "Baltic_Commerce_Analysis.xlsx", channels, markets, logistics, experiment)
    finally:
        conn.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=Path(__file__).resolve().parents[1])
    run(parser.parse_args().root)
